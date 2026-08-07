"""api/routes/reports.py"""
from email.message import EmailMessage
from email.utils import formataddr, formatdate, make_msgid, parseaddr
import html
import logging
import smtplib
import ssl
import threading
import time
import re
import calendar
from io import BytesIO
from zoneinfo import ZoneInfo
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy.orm import selectinload
from sqlalchemy import func
from datetime import date, datetime, timedelta
from core.database import get_db, SessionLocal
from core.security import get_current_user
from core.settings_store import load_settings, save_settings
from api.schemas import ReportEmailRequest, ReportEmailTestRequest
from models.sales import Sale, SaleItem
from models.expense import Expense
from models.purchase import Purchase
from models.product import Product
from models.user import User
from services.document_workflow import ACTIVE_PURCHASE_STATUSES, ACTIVE_SALE_STATUSES
from services.money import ZERO, quantize_money
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

router = APIRouter()
_SCHEDULER_STARTED = False
logger = logging.getLogger("secureerp.reports")


def _load_settings():
    return load_settings(include_secrets=True)


def _save_settings(settings: dict):
    save_settings(settings)


def _period_key(frequency: str, now: datetime, settings: dict):
    if frequency == "daily":
        return now.strftime("%Y-%m-%d")
    if frequency == "weekly":
        return f"{now.isocalendar().year}-W{now.isocalendar().week:02d}"
    if frequency == "monthly":
        return now.strftime("%Y-%m")
    if frequency == "yearly":
        return now.strftime("%Y")
    return now.strftime("%Y-%m-%d")


def _is_schedule_due(settings: dict, now: datetime):
    if not settings.get("report_email_enabled"):
        return False
    if not settings.get("report_email_recipients"):
        return False
    scheduled_time = settings.get("report_schedule_time") or "20:00"
    try:
        hour, minute = [int(part) for part in scheduled_time.split(":")[:2]]
    except Exception:
        hour, minute = 20, 0
    if (now.hour, now.minute) < (hour, minute):
        return False

    frequency = settings.get("report_schedule_frequency") or "daily"
    if frequency == "weekly" and now.isoweekday() != int(settings.get("report_schedule_day_of_week") or 1):
        return False
    configured_day = int(settings.get("report_schedule_day_of_month") or 1)
    effective_day = min(configured_day, calendar.monthrange(now.year, now.month)[1])
    if frequency == "monthly" and now.day != effective_day:
        return False
    if frequency == "yearly":
        if now.month != int(settings.get("report_schedule_month") or 1):
            return False
        if now.day != effective_day:
            return False

    last_key = settings.get("report_schedule_last_sent_key") or ""
    return last_key != _period_key(frequency, now, settings)


def _date_range(period_type: str, start_date: str = None, end_date: str = None):
    today = date.today()
    if period_type == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="start_date et end_date sont requis")
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="Dates invalides. Format attendu: AAAA-MM-JJ") from exc
    elif period_type == "daily":
        start = end = today
    elif period_type == "weekly":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
    elif period_type == "monthly":
        start = today.replace(day=1)
        next_month = start.replace(year=start.year + 1, month=1) if start.month == 12 else start.replace(month=start.month + 1)
        end = next_month - timedelta(days=1)
    elif period_type == "yearly":
        start = today.replace(month=1, day=1)
        end = today.replace(month=12, day=31)
    else:
        raise HTTPException(status_code=400, detail="period_type invalide")
    if start > end:
        raise HTTPException(status_code=400, detail="La date debut doit etre avant la date fin")
    return start, end


def _between_dates(column, start: date, end: date):
    return func.date(column) >= start, func.date(column) <= end


def _csv_emails(value: str):
    candidates = [item.strip() for item in (value or "").replace(";", ",").split(",") if item.strip()]
    result = []
    seen = set()
    for candidate in candidates:
        _, address = parseaddr(candidate)
        address = address.strip().lower()
        if not re.fullmatch(r"[^\s@,]+@[^\s@,]+\.[^\s@,]+", address):
            raise HTTPException(status_code=422, detail=f"Adresse email invalide: {candidate}")
        if address not in seen:
            seen.add(address)
            result.append(address)
    return result


def _safe_header(value: str, fallback: str = "") -> str:
    return " ".join(str(value or fallback).replace("\x00", "").splitlines()).strip()


def _money(value):
    return f"{quantize_money(value or 0):,.2f}".replace(",", " ")


def _build_summary(db: Session, start: date, end: date):
    sale_filters = [
        Sale.doc_type == "invoice",
        Sale.status.in_(ACTIVE_SALE_STATUSES),
        *_between_dates(Sale.date_time, start, end),
    ]
    revenue = db.query(func.sum(Sale.total_amount)).filter(*sale_filters).scalar() or ZERO
    paid = db.query(func.sum(Sale.paid_amount)).filter(*sale_filters).scalar() or ZERO
    sale_count = db.query(func.count(Sale.id)).filter(*sale_filters).scalar() or 0
    cogs = db.query(func.sum(SaleItem.purchase_price * SaleItem.quantity)).join(Sale).filter(*sale_filters).scalar() or ZERO
    expenses = db.query(func.sum(Expense.amount)).filter(*_between_dates(Expense.date, start, end)).scalar() or ZERO
    purchases = db.query(func.sum(Purchase.total_amount)).filter(
        Purchase.status.in_(ACTIVE_PURCHASE_STATUSES),
        *_between_dates(Purchase.date_time, start, end),
    ).scalar() or ZERO
    purchase_count = db.query(func.count(Purchase.id)).filter(
        Purchase.status.in_(ACTIVE_PURCHASE_STATUSES),
        *_between_dates(Purchase.date_time, start, end),
    ).scalar() or 0

    gross_profit = revenue - cogs
    net_profit = gross_profit - expenses
    margin = (gross_profit / revenue * 100) if revenue > 0 else 0
    return {
        "revenue": quantize_money(revenue),
        "paid": quantize_money(paid),
        "unpaid": quantize_money(max(revenue - paid, 0)),
        "sale_count": sale_count,
        "cogs": quantize_money(cogs),
        "gross_profit": quantize_money(gross_profit),
        "expenses": quantize_money(expenses),
        "purchases": quantize_money(purchases),
        "purchase_count": purchase_count,
        "net_profit": quantize_money(net_profit),
        "margin_pct": round(margin, 2),
    }


def _sales_by_category_range(db: Session, start: date, end: date):
    from models.product import Category
    rows = db.query(
        Category.name.label("category"),
        func.sum(SaleItem.line_total).label("total"),
        func.count(SaleItem.id).label("count"),
    ).join(Product, SaleItem.product_id == Product.id
    ).join(Category, Product.category_id == Category.id
    ).join(Sale, SaleItem.sale_id == Sale.id
    ).filter(
        Sale.doc_type == "invoice", Sale.status.in_(ACTIVE_SALE_STATUSES),
        *_between_dates(Sale.date_time, start, end),
    ).group_by(Category.name).order_by(func.sum(SaleItem.line_total).desc()).limit(10).all()
    return [{"category": r.category, "total": quantize_money(r.total or 0), "count": r.count} for r in rows]


def _sales_by_user_range(db: Session, start: date, end: date):
    display_name = func.coalesce(func.nullif(User.full_name, ""), func.nullif(User.username, ""), "Utilisateur inconnu")
    rows = db.query(
        display_name.label("user_name"),
        func.count(Sale.id).label("invoice_count"),
        func.sum(Sale.total_amount).label("revenue"),
        func.sum(Sale.paid_amount).label("paid"),
    ).outerjoin(User, User.id == Sale.created_by).filter(
        Sale.doc_type == "invoice",
        Sale.status.in_(ACTIVE_SALE_STATUSES),
        *_between_dates(Sale.date_time, start, end),
    ).group_by(display_name).order_by(func.sum(Sale.total_amount).desc()).all()
    return [{
        "user_name": row.user_name or "Utilisateur inconnu",
        "invoice_count": int(row.invoice_count or 0),
        "revenue": quantize_money(row.revenue or 0),
        "paid": quantize_money(row.paid or 0),
    } for row in rows]


def _stock_summary(db: Session):
    rows = db.query(Product).filter(Product.is_active == 1, Product.product_type == "product").all()
    total_value = sum((p.stock_quantity or 0) * (p.purchase_price or 0) for p in rows)
    low_stock = [p for p in rows if (p.stock_quantity or 0) <= (p.min_stock or 0)]
    return {"total_value": quantize_money(total_value), "low_stock_count": len(low_stock), "products_count": len(rows)}


def _cash_summary(db: Session, start: date, end: date):
    from models.cash import CashTransaction
    cash_in = db.query(func.sum(CashTransaction.amount)).filter(
        CashTransaction.direction == "in",
        *_between_dates(CashTransaction.created_at, start, end),
    ).scalar() or ZERO
    cash_out = db.query(func.sum(CashTransaction.amount)).filter(
        CashTransaction.direction == "out",
        *_between_dates(CashTransaction.created_at, start, end),
    ).scalar() or ZERO
    return {"cash_in": quantize_money(cash_in), "cash_out": quantize_money(cash_out), "net_cash": quantize_money(cash_in - cash_out)}


def _period_comparison_range(start: date, end: date):
    span_days = (end - start).days + 1
    previous_end = start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=span_days - 1)
    return previous_start, previous_end


def _variation(current: float, previous: float):
    current = current or 0
    previous = previous or 0
    if previous == 0:
        return 100.0 if current > 0 else 0.0
    return round(((current - previous) / previous) * 100, 2)


def _report_timeseries(db: Session, start: date, end: date):
    use_month_bucket = (end - start).days > 62
    bucket_expr = func.strftime("%Y-%m", Sale.date_time) if use_month_bucket else func.date(Sale.date_time)
    expense_bucket = func.strftime("%Y-%m", Expense.date) if use_month_bucket else func.date(Expense.date)
    purchase_bucket = func.strftime("%Y-%m", Purchase.date_time) if use_month_bucket else func.date(Purchase.date_time)

    labels = []
    cursor = start
    if use_month_bucket:
        cursor = start.replace(day=1)
        while cursor <= end:
            labels.append(cursor.strftime("%Y-%m"))
            cursor = cursor.replace(year=cursor.year + 1, month=1) if cursor.month == 12 else cursor.replace(month=cursor.month + 1)
    else:
        while cursor <= end:
            labels.append(cursor.isoformat())
            cursor += timedelta(days=1)

    series = {label: {"period": label, "revenue": ZERO, "expenses": ZERO, "purchases": ZERO} for label in labels}

    sales_rows = db.query(bucket_expr.label("period"), func.sum(Sale.total_amount).label("total")).filter(
        Sale.doc_type == "invoice",
        Sale.status.in_(ACTIVE_SALE_STATUSES),
        *_between_dates(Sale.date_time, start, end),
    ).group_by(bucket_expr).all()
    for row in sales_rows:
        if row.period in series:
            series[row.period]["revenue"] = quantize_money(row.total or 0)

    expense_rows = db.query(expense_bucket.label("period"), func.sum(Expense.amount).label("total")).filter(
        *_between_dates(Expense.date, start, end),
    ).group_by(expense_bucket).all()
    for row in expense_rows:
        if row.period in series:
            series[row.period]["expenses"] = quantize_money(row.total or 0)

    purchase_rows = db.query(purchase_bucket.label("period"), func.sum(Purchase.total_amount).label("total")).filter(
        Purchase.status.in_(ACTIVE_PURCHASE_STATUSES),
        *_between_dates(Purchase.date_time, start, end),
    ).group_by(purchase_bucket).all()
    for row in purchase_rows:
        if row.period in series:
            series[row.period]["purchases"] = quantize_money(row.total or 0)

    return list(series.values())


def _stock_value_detail(db: Session, limit: int = 120):
    from models.product import Category
    rows = db.query(
        Product.id, Product.name, Product.code,
        Product.stock_quantity, Product.min_stock, Product.purchase_price, Product.sale_price,
        Category.name.label("category"),
    ).join(Category, Product.category_id == Category.id, isouter=True
    ).filter(Product.is_active == 1, Product.product_type == "product"
    ).order_by(func.coalesce(Category.name, "zzz"), Product.name).limit(limit).all()
    total_value = db.query(func.sum(Product.stock_quantity * Product.purchase_price)).filter(
        Product.is_active == 1,
        Product.product_type == "product",
    ).scalar() or ZERO
    low_stock_count = db.query(func.count(Product.id)).filter(
        Product.is_active == 1,
        Product.product_type == "product",
        Product.stock_quantity <= Product.min_stock,
    ).scalar() or 0
    return {
        "total_value": quantize_money(total_value),
        "low_stock_count": low_stock_count,
        "products": [{
            "id": r.id,
            "name": r.name,
            "code": r.code or "",
            "category": r.category or "-",
            "stock": r.stock_quantity or 0,
            "min_stock": r.min_stock or 0,
            "purchase_price": r.purchase_price or 0,
            "sale_price": r.sale_price or 0,
            "value": quantize_money((r.stock_quantity or 0) * (r.purchase_price or 0)),
            "is_low": (r.stock_quantity or 0) <= (r.min_stock or 0),
        } for r in rows],
    }


def _top_items(db: Session, start: date, end: date, limit: int = 12):
    rows = db.query(
        SaleItem.product_id,
        func.coalesce(func.max(Product.name), func.max(SaleItem.description), "—").label("name"),
        func.coalesce(func.max(Product.product_type), "product").label("product_type"),
        func.sum(SaleItem.quantity).label("quantity"),
        func.sum(SaleItem.line_total).label("revenue"),
        func.sum(SaleItem.quantity * SaleItem.purchase_price).label("cost"),
    ).join(Sale, Sale.id == SaleItem.sale_id
    ).join(Product, Product.id == SaleItem.product_id, isouter=True
    ).filter(
        Sale.doc_type == "invoice",
        Sale.status.in_(ACTIVE_SALE_STATUSES),
        *_between_dates(Sale.date_time, start, end),
    ).group_by(SaleItem.product_id).order_by(func.sum(SaleItem.quantity).desc()).limit(limit).all()
    return [{
        "product_id": row.product_id,
        "name": row.name,
        "product_type": row.product_type,
        "quantity": float(row.quantity or 0),
        "revenue": quantize_money(row.revenue or 0),
        "gross_profit": quantize_money((row.revenue or 0) - (row.cost or 0)),
    } for row in rows]


def _hourly_performance(db: Session, start: date, end: date):
    hour_expr = func.strftime("%H", Sale.date_time)
    rows = db.query(
        hour_expr.label("hour"),
        func.count(Sale.id).label("sales_count"),
        func.sum(Sale.total_amount).label("revenue"),
    ).filter(
        Sale.doc_type == "invoice",
        Sale.status.in_(ACTIVE_SALE_STATUSES),
        *_between_dates(Sale.date_time, start, end),
    ).group_by(hour_expr).all()
    by_hour = {int(row.hour): row for row in rows if row.hour is not None}
    return [{
        "hour": f"{hour:02d}:00",
        "sales_count": int(by_hour[hour].sales_count or 0) if hour in by_hour else 0,
        "revenue": quantize_money(by_hour[hour].revenue or 0) if hour in by_hour else quantize_money(0),
    } for hour in range(7, 23)]


def _dormant_products(db: Session, inactivity_days: int = 60, limit: int = 30):
    cutoff = date.today() - timedelta(days=inactivity_days)
    last_sale = db.query(
        SaleItem.product_id.label("product_id"),
        func.max(Sale.date_time).label("last_sale_at"),
    ).join(Sale, Sale.id == SaleItem.sale_id).filter(
        Sale.doc_type == "invoice",
        Sale.status.in_(ACTIVE_SALE_STATUSES),
    ).group_by(SaleItem.product_id).subquery()
    rows = db.query(Product, last_sale.c.last_sale_at).outerjoin(
        last_sale, last_sale.c.product_id == Product.id,
    ).filter(
        Product.is_active == 1,
        Product.product_type == "product",
        Product.stock_quantity > 0,
        (last_sale.c.last_sale_at.is_(None)) | (func.date(last_sale.c.last_sale_at) < cutoff),
    ).order_by(last_sale.c.last_sale_at.asc()).limit(limit).all()
    return [{
        "product_id": product.id,
        "code": product.code or "",
        "name": product.name,
        "stock": float(product.stock_quantity or 0),
        "stock_value": quantize_money((product.stock_quantity or 0) * (product.purchase_price or 0)),
        "last_sale_at": last_at.isoformat() if last_at else None,
        "inactive_days": (date.today() - last_at.date()).days if last_at else None,
    } for product, last_at in rows]


def _business_comparisons(db: Session):
    today = date.today()
    yesterday = today - timedelta(days=1)
    month_start = today.replace(day=1)
    previous_month_end = month_start - timedelta(days=1)
    previous_month_start = previous_month_end.replace(day=1)
    today_data = _build_summary(db, today, today)
    yesterday_data = _build_summary(db, yesterday, yesterday)
    month_data = _build_summary(db, month_start, today)
    comparable_previous_end = min(
        previous_month_start + timedelta(days=(today - month_start).days),
        previous_month_end,
    )
    previous_month_data = _build_summary(db, previous_month_start, comparable_previous_end)
    return {
        "today": today_data,
        "yesterday": yesterday_data,
        "today_vs_yesterday": {
            key: _variation(today_data[key], yesterday_data[key])
            for key in ("revenue", "sale_count", "net_profit")
        },
        "month": month_data,
        "previous_month": previous_month_data,
        "month_vs_previous": {
            key: _variation(month_data[key], previous_month_data[key])
            for key in ("revenue", "sale_count", "net_profit")
        },
    }


def _render_report_html(settings: dict, req: ReportEmailRequest, start: date, end: date, data: dict):
    company = html.escape(settings.get("name") or settings.get("store_name") or "Maktaba Print")
    currency = html.escape(settings.get("currency") or "MAD")
    summary = data["summary"]
    trend = data.get("trend", {})
    generated_at = datetime.now(ZoneInfo(settings.get("report_schedule_timezone") or "Africa/Casablanca"))

    def kpi(label, value, color, variation=None):
        variation_html = ""
        if variation is not None:
            sign = "+" if variation > 0 else ""
            trend_color = "#16a34a" if variation >= 0 else "#dc2626"
            variation_html = f'<div style="margin-top:5px;font-size:12px;color:{trend_color};font-weight:700">{sign}{variation:.1f}% vs periode precedente</div>'
        return f'''<td width="33.33%" style="padding:6px;vertical-align:top">
          <div style="border:1px solid #dbe7f5;border-top:4px solid {color};border-radius:12px;padding:14px;background:#ffffff;min-height:82px">
            <div style="font-size:12px;color:#64748b;margin-bottom:7px">{html.escape(label)}</div>
            <div style="font-size:20px;color:#102a43;font-weight:800;white-space:nowrap">{_money(value)} {currency}</div>{variation_html}
          </div></td>'''

    primary_kpis = "".join([
        kpi("Chiffre d'affaires", summary["revenue"], "#1677ff", trend.get("revenue")),
        kpi("Encaisse", summary["paid"], "#16a34a", trend.get("paid")),
        kpi("Reste a encaisser", summary["unpaid"], "#f59e0b"),
    ])
    secondary_kpis = []
    if req.include_profit:
        secondary_kpis.append(kpi("Benefice net", summary["net_profit"], "#16a34a" if summary["net_profit"] >= 0 else "#dc2626", trend.get("net_profit")))
    if req.include_expenses:
        secondary_kpis.append(kpi("Depenses", summary["expenses"], "#ef4444", trend.get("expenses")))
    if req.include_purchases:
        secondary_kpis.append(kpi("Achats", summary["purchases"], "#8b5cf6", trend.get("purchases")))

    categories = "".join(
        f"<tr><td style='padding:10px;border-bottom:1px solid #e5edf6'>{html.escape(r['category'])}</td><td align='center' style='padding:10px;border-bottom:1px solid #e5edf6'>{r['count']}</td><td align='right' style='padding:10px;border-bottom:1px solid #e5edf6;font-weight:700'>{_money(r['total'])} {currency}</td></tr>"
        for r in data.get("categories", [])
    ) or "<tr><td colspan='3' style='padding:14px;color:#64748b;text-align:center'>Aucune vente sur cette periode</td></tr>"
    top_items = "".join(
        f"<tr><td style='padding:10px;border-bottom:1px solid #e5edf6'>{html.escape(str(item['name']))}<div style='font-size:11px;color:#64748b'>{'Service' if item['product_type'] == 'service' else 'Produit'}</div></td><td align='center' style='padding:10px;border-bottom:1px solid #e5edf6'>{item['quantity']:g}</td><td align='right' style='padding:10px;border-bottom:1px solid #e5edf6;font-weight:700'>{_money(item['revenue'])} {currency}</td></tr>"
        for item in data.get("top_items", [])[:8]
    ) or "<tr><td colspan='3' style='padding:14px;color:#64748b;text-align:center'>Aucun article vendu</td></tr>"
    users = "".join(
        f"<tr><td style='padding:10px;border-bottom:1px solid #e5edf6;font-weight:700'>{html.escape(str(row['user_name']))}</td><td align='center' style='padding:10px;border-bottom:1px solid #e5edf6'>{row['invoice_count']}</td><td align='right' style='padding:10px;border-bottom:1px solid #e5edf6'>{_money(row['paid'])} {currency}</td><td align='right' style='padding:10px;border-bottom:1px solid #e5edf6;font-weight:700'>{_money(row['revenue'])} {currency}</td></tr>"
        for row in data.get("users", [])
    ) or "<tr><td colspan='4' style='padding:14px;color:#64748b;text-align:center'>Aucune facture utilisateur</td></tr>"
    section_title = "font-size:18px;color:#102a43;margin:26px 0 10px;font-weight:800"
    table_header = "padding:10px;background:#1677ff;color:#ffffff;font-size:12px;text-transform:uppercase"
    return f"""<!doctype html><html><head><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f3f7fc;font-family:Arial,Helvetica,sans-serif;color:#102a43">
<div style="display:none;max-height:0;overflow:hidden">Rapport du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')} - CA {_money(summary['revenue'])} {currency}</div>
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f3f7fc"><tr><td align="center" style="padding:22px 10px">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:720px;background:#ffffff;border-radius:18px;overflow:hidden;border:1px solid #dbe7f5">
<tr><td style="padding:25px 28px;background:linear-gradient(135deg,#0b5ed7,#1677ff);color:#ffffff">
  <div style="font-size:13px;opacity:.88;letter-spacing:.7px">RAPPORT DE GESTION</div>
  <div style="font-size:26px;font-weight:800;margin-top:6px">{company}</div>
  <div style="font-size:14px;margin-top:8px;opacity:.92">Du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}</div>
</td></tr>
<tr><td style="padding:22px 20px 8px">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>{primary_kpis}</tr></table>
  {f'<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>{"".join(secondary_kpis)}</tr></table>' if secondary_kpis else ''}
  <div style="margin:12px 6px 0;padding:13px 15px;border-radius:10px;background:#eff6ff;color:#334e68;font-size:13px">
    <strong>{summary['sale_count']}</strong> ventes &nbsp;•&nbsp; Marge brute <strong>{summary['margin_pct']:.2f}%</strong> &nbsp;•&nbsp; <strong>{summary['purchase_count']}</strong> achats
  </div>
  {f'''<div style="{section_title}">Ventes par categorie</div><table width="100%" cellspacing="0" cellpadding="0" style="border:1px solid #dbe7f5;border-radius:10px;overflow:hidden"><tr><th align="left" style="{table_header}">Categorie</th><th style="{table_header}">Lignes</th><th align="right" style="{table_header}">Chiffre d'affaires</th></tr>{categories}</table>''' if req.include_sales_by_category else ''}
  <div style="{section_title}">Meilleures ventes</div><table width="100%" cellspacing="0" cellpadding="0" style="border:1px solid #dbe7f5;border-radius:10px;overflow:hidden"><tr><th align="left" style="{table_header}">Article</th><th style="{table_header}">Quantite</th><th align="right" style="{table_header}">CA</th></tr>{top_items}</table>
  <div style="{section_title}">Performance par utilisateur</div><table width="100%" cellspacing="0" cellpadding="0" style="border:1px solid #dbe7f5;border-radius:10px;overflow:hidden"><tr><th align="left" style="{table_header}">Utilisateur</th><th style="{table_header}">Factures</th><th align="right" style="{table_header}">Encaisse</th><th align="right" style="{table_header}">CA</th></tr>{users}</table>
  {f'''<div style="{section_title}">Stock</div><table width="100%" cellspacing="0" cellpadding="0"><tr><td style="padding:14px;background:#f8fafc;border-radius:10px">Valeur du stock<br><strong style="font-size:18px">{_money(data['stock']['total_value'])} {currency}</strong></td><td width="10"></td><td style="padding:14px;background:{'#fff7ed' if data['stock']['low_stock_count'] else '#f0fdf4'};border-radius:10px">Stock faible<br><strong style="font-size:18px;color:{'#ea580c' if data['stock']['low_stock_count'] else '#16a34a'}">{data['stock']['low_stock_count']} produit(s)</strong></td></tr></table>''' if req.include_stock_value else ''}
  {f'''<div style="{section_title}">Caisse</div><table width="100%" cellspacing="0" cellpadding="0" style="background:#f8fafc;border-radius:10px"><tr><td style="padding:14px">Entrees<br><strong style="color:#16a34a">{_money(data['cash']['cash_in'])} {currency}</strong></td><td style="padding:14px">Sorties<br><strong style="color:#dc2626">{_money(data['cash']['cash_out'])} {currency}</strong></td><td style="padding:14px">Solde net<br><strong>{_money(data['cash']['net_cash'])} {currency}</strong></td></tr></table>''' if req.include_cash else ''}
</td></tr>
<tr><td style="padding:20px 28px;text-align:center;color:#7b8fa3;font-size:11px;border-top:1px solid #e5edf6">Rapport genere le {generated_at.strftime('%d/%m/%Y a %H:%M')} • {company}</td></tr>
</table></td></tr></table></body></html>"""


def _render_report_text(settings: dict, start: date, end: date, data: dict) -> str:
    summary = data["summary"]
    currency = settings.get("currency") or "MAD"
    return "\n".join([
        str(settings.get("name") or settings.get("store_name") or "Maktaba Print"),
        f"Rapport du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}",
        "",
        f"Chiffre d'affaires: {_money(summary['revenue'])} {currency}",
        f"Encaisse: {_money(summary['paid'])} {currency}",
        f"Reste a encaisser: {_money(summary['unpaid'])} {currency}",
        f"Benefice net: {_money(summary['net_profit'])} {currency}",
        f"Ventes: {summary['sale_count']}",
    ])


def _build_email_report_data(db: Session, start: date, end: date, req: ReportEmailRequest) -> dict:
    previous_start, previous_end = _period_comparison_range(start, end)
    summary = _build_summary(db, start, end)
    previous = _build_summary(db, previous_start, previous_end)
    return {
        "summary": summary,
        "previous": previous,
        "trend": {key: _variation(summary[key], previous[key]) for key in ("revenue", "paid", "expenses", "purchases", "net_profit")},
        "categories": _sales_by_category_range(db, start, end) if req.include_sales_by_category else [],
        "top_items": _top_items(db, start, end, limit=8),
        "users": _sales_by_user_range(db, start, end),
        "stock": _stock_summary(db) if req.include_stock_value else {"total_value": 0, "products_count": 0, "low_stock_count": 0},
        "cash": _cash_summary(db, start, end) if req.include_cash else {"cash_in": 0, "cash_out": 0, "net_cash": 0},
    }


def _safe_sheet_title(value: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", str(value or "Facture")).strip() or "Facture"
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate.lower() in used:
        marker = f"-{suffix}"
        candidate = f"{base[:31-len(marker)]}{marker}"
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _build_invoices_workbook(db: Session, settings: dict, start: date, end: date, data: dict) -> tuple[bytes, str]:
    invoices = db.query(Sale).options(
        selectinload(Sale.client),
        selectinload(Sale.creator),
        selectinload(Sale.items).selectinload(SaleItem.product),
    ).filter(
        Sale.doc_type == "invoice",
        Sale.status.in_(ACTIVE_SALE_STATUSES),
        *_between_dates(Sale.date_time, start, end),
    ).order_by(Sale.date_time, Sale.id).all()

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resume"
    blue = "1677FF"
    navy = "102A43"
    pale_blue = "EAF3FF"
    pale_gray = "F4F7FA"
    white = "FFFFFF"
    green = "16A34A"
    red = "DC2626"
    thin = Side(style="thin", color="D9E4EF")
    currency = str(settings.get("currency") or "MAD")

    summary_sheet.merge_cells("A1:F1")
    summary_sheet["A1"] = str(settings.get("name") or settings.get("store_name") or "LIBRARY SABRI")
    summary_sheet["A1"].font = Font(size=18, bold=True, color=white)
    summary_sheet["A1"].fill = PatternFill("solid", fgColor=blue)
    summary_sheet["A1"].alignment = Alignment(horizontal="center")
    summary_sheet.row_dimensions[1].height = 30
    summary_sheet.merge_cells("A2:F2")
    summary_sheet["A2"] = f"Rapport des factures du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}"
    summary_sheet["A2"].font = Font(bold=True, color=navy)
    summary_sheet["A2"].alignment = Alignment(horizontal="center")
    summary_sheet.append([])
    summary_sheet.append(["Indicateur", "Valeur"])
    metrics = [
        ("Nombre de factures", len(invoices)),
        ("Chiffre d'affaires", float(data["summary"]["revenue"])),
        ("Montant encaisse", float(data["summary"]["paid"])),
        ("Reste a encaisser", float(data["summary"]["unpaid"])),
        ("Benefice net", float(data["summary"]["net_profit"])),
    ]
    for label, value in metrics:
        summary_sheet.append([label, value])
    for cell in summary_sheet[4]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color=white)
    for row in range(6, 10):
        summary_sheet.cell(row, 2).number_format = f'#,##0.00 "{currency}"'

    users_sheet = workbook.create_sheet("Utilisateurs")
    users_sheet.append(["Utilisateur", "Nombre de factures", "Chiffre d'affaires", "Montant encaisse", "Reste"])
    for cell in users_sheet[1]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(horizontal="center")
    for user_row in data.get("users", []):
        users_sheet.append([
            user_row["user_name"], user_row["invoice_count"], float(user_row["revenue"]),
            float(user_row["paid"]), float(user_row["revenue"] - user_row["paid"]),
        ])
    for row in range(2, users_sheet.max_row + 1):
        for column in (3, 4, 5):
            users_sheet.cell(row, column).number_format = f'#,##0.00 "{currency}"'
    users_sheet.freeze_panes = "A2"
    users_sheet.auto_filter.ref = f"A1:E{max(1, users_sheet.max_row)}"
    for column, width in enumerate([30, 20, 22, 22, 18], 1):
        users_sheet.column_dimensions[get_column_letter(column)].width = width

    index_sheet = workbook.create_sheet("Index factures")
    headers = ["N° facture", "Date", "Client", "Utilisateur", "Statut", "Paiement", "Total", "Paye", "Reste", "Feuille"]
    index_sheet.append(headers)
    for cell in index_sheet[1]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(horizontal="center")

    used_titles = {"resume", "utilisateurs", "index factures"}
    for invoice in invoices:
        title = _safe_sheet_title(invoice.number or f"Facture-{invoice.id}", used_titles)
        client_name = invoice.client.name if invoice.client else "Client comptoir"
        creator_name = (invoice.creator.full_name or invoice.creator.username) if invoice.creator else "Utilisateur inconnu"
        balance = max(float(invoice.total_amount or 0) - float(invoice.paid_amount or 0), 0)
        index_sheet.append([
            invoice.number or f"#{invoice.id}", invoice.date_time, client_name, creator_name, invoice.status,
            invoice.payment_mode or "", float(invoice.total_amount or 0), float(invoice.paid_amount or 0), balance, title,
        ])
        invoice_sheet = workbook.create_sheet(title)
        invoice_sheet.sheet_view.showGridLines = False
        invoice_sheet.merge_cells("A1:F1")
        invoice_sheet["A1"] = f"FACTURE {invoice.number or invoice.id}"
        invoice_sheet["A1"].fill = PatternFill("solid", fgColor=blue)
        invoice_sheet["A1"].font = Font(size=17, bold=True, color=white)
        invoice_sheet["A1"].alignment = Alignment(horizontal="center")
        invoice_sheet.row_dimensions[1].height = 29
        invoice_sheet["A3"] = "Client"
        invoice_sheet["B3"] = client_name
        invoice_sheet["D3"] = "Date"
        invoice_sheet["E3"] = invoice.date_time
        invoice_sheet["A4"] = "Mode paiement"
        invoice_sheet["B4"] = invoice.payment_mode or ""
        invoice_sheet["D4"] = "Utilisateur"
        invoice_sheet["E4"] = creator_name
        invoice_sheet["A5"] = "Statut"
        invoice_sheet["B5"] = invoice.status
        for coordinate in ("A3", "D3", "A4", "D4", "A5"):
            invoice_sheet[coordinate].font = Font(bold=True, color=navy)

        row = 6
        line_headers = ["Designation", "Quantite", "Prix unitaire", "Remise %", "TVA %", "Total"]
        for column, label in enumerate(line_headers, 1):
            cell = invoice_sheet.cell(row, column, label)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(bold=True, color=white)
            cell.alignment = Alignment(horizontal="center")
        for item in invoice.items:
            row += 1
            invoice_sheet.append([
                item.description or (item.product.name if item.product else "Article"),
                float(item.quantity or 0), float(item.unit_price or 0), float(item.discount or 0),
                float(item.tax_rate or 0), float(item.total_amount or item.line_total or 0),
            ])
            invoice_sheet.cell(row, 2).number_format = "0.####"
            for column in (3, 6):
                invoice_sheet.cell(row, column).number_format = f'#,##0.00 "{currency}"'
            for column in (4, 5):
                invoice_sheet.cell(row, column).number_format = '0.00"%"'
        row += 2
        totals = [
            ("Sous-total", invoice.subtotal), ("Remise", invoice.discount_amount),
            ("TVA", invoice.tax_amount), ("TOTAL", invoice.total_amount),
            ("Paye", invoice.paid_amount), ("Reste", balance),
        ]
        for label, value in totals:
            invoice_sheet.cell(row, 5, label)
            invoice_sheet.cell(row, 6, float(value or 0))
            invoice_sheet.cell(row, 5).font = Font(bold=True, color=navy)
            invoice_sheet.cell(row, 6).font = Font(bold=True, color=green if label == "Paye" else red if label == "Reste" and float(value or 0) > 0 else navy)
            invoice_sheet.cell(row, 6).number_format = f'#,##0.00 "{currency}"'
            row += 1
        invoice_sheet.freeze_panes = "A7"
        invoice_sheet.auto_filter.ref = f"A6:F{max(6, row - 8)}"
        widths = [42, 12, 17, 13, 11, 18]
        for column, width in enumerate(widths, 1):
            invoice_sheet.column_dimensions[get_column_letter(column)].width = width
        for cells in invoice_sheet.iter_rows(min_row=3, max_row=row - 1, min_col=1, max_col=6):
            for cell in cells:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in range(2, index_sheet.max_row + 1):
        index_sheet.cell(row, 2).number_format = "dd/mm/yyyy hh:mm"
        for column in (7, 8, 9):
            index_sheet.cell(row, column).number_format = f'#,##0.00 "{currency}"'
        sheet_cell = index_sheet.cell(row, 10)
        sheet_cell.hyperlink = f"#'{sheet_cell.value}'!A1"
        sheet_cell.style = "Hyperlink"
        if row % 2 == 0:
            for cell in index_sheet[row]:
                cell.fill = PatternFill("solid", fgColor=pale_gray)
    index_sheet.freeze_panes = "A2"
    index_sheet.auto_filter.ref = f"A1:J{max(1, index_sheet.max_row)}"
    for column, width in enumerate([18, 18, 28, 25, 18, 18, 16, 16, 16, 20], 1):
        index_sheet.column_dimensions[get_column_letter(column)].width = width
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 22

    stream = BytesIO()
    workbook.save(stream)
    filename = f"factures-{start.isoformat()}-{end.isoformat()}.xlsx"
    return stream.getvalue(), filename


def _send_report_from_settings(db: Session, settings: dict, period_type: str = None):
    req = ReportEmailRequest(
        period_type=period_type or settings.get("report_schedule_frequency") or "daily",
        recipients=settings.get("report_email_recipients", ""),
        include_profit=bool(settings.get("report_email_include_profit", True)),
        include_sales_by_category=bool(settings.get("report_email_include_sales_by_category", True)),
        include_stock_value=bool(settings.get("report_email_include_stock_value", True)),
        include_cash=bool(settings.get("report_email_include_cash", True)),
        include_expenses=bool(settings.get("report_email_include_expenses", True)),
        include_purchases=bool(settings.get("report_email_include_purchases", True)),
    )
    start, end = _date_range(req.period_type)
    data = _build_email_report_data(db, start, end, req)
    recipients = _csv_emails(req.recipients)
    subject = f"{settings.get('report_email_subject_prefix') or 'Rapport Maktaba Print'} - {start.isoformat()} / {end.isoformat()}"
    html_body = _render_report_html(settings, req, start, end, data)
    excel_bytes, excel_name = _build_invoices_workbook(db, settings, start, end, data)
    _send_email(settings, recipients, subject, html_body, _render_report_text(settings, start, end, data), [(excel_name, excel_bytes)])
    return {"recipients": recipients, "start": start, "end": end, "subject": subject}


def _scheduler_loop():
    while True:
        try:
            settings = _load_settings()
            tz = ZoneInfo(settings.get("report_schedule_timezone") or "Africa/Casablanca")
            now = datetime.now(tz)
            if _is_schedule_due(settings, now):
                db = SessionLocal()
                try:
                    _send_report_from_settings(db, settings)
                    settings["report_schedule_last_sent_at"] = now.isoformat()
                    settings["report_schedule_last_sent_key"] = _period_key(settings.get("report_schedule_frequency") or "daily", now, settings)
                    settings["report_schedule_last_status"] = "success"
                    settings["report_schedule_last_error"] = ""
                    _save_settings(settings)
                except Exception as exc:
                    settings["report_schedule_last_status"] = "error"
                    settings["report_schedule_last_error"] = _smtp_error_message(exc)
                    settings["report_schedule_last_attempt_at"] = now.isoformat()
                    _save_settings(settings)
                    logger.exception("Scheduled report email failed")
                finally:
                    db.close()
        except Exception:
            logger.exception("Report email scheduler iteration failed")
        time.sleep(60)


def start_report_email_scheduler():
    global _SCHEDULER_STARTED
    if _SCHEDULER_STARTED:
        return
    _SCHEDULER_STARTED = True
    thread = threading.Thread(target=_scheduler_loop, name="report-email-scheduler", daemon=True)
    thread.start()


def _send_email(settings: dict, to_emails: list[str], subject: str, html_body: str, text_body: str = "", attachments: list[tuple[str, bytes]] | None = None):
    if not settings.get("smtp_host") or not settings.get("smtp_from_email"):
        raise HTTPException(status_code=400, detail="Configuration SMTP incomplete")
    if not to_emails:
        raise HTTPException(status_code=400, detail="Aucun destinataire configure")

    from_email = _csv_emails(settings.get("smtp_from_email", ""))
    if len(from_email) != 1:
        raise HTTPException(status_code=422, detail="Adresse d'expediteur SMTP invalide")
    smtp_username = _csv_emails(settings.get("smtp_username", "")) if "@" in str(settings.get("smtp_username", "")) else []
    if str(settings.get("smtp_host", "")).lower() == "smtp.gmail.com" and smtp_username:
        # Gmail signs the authenticated account. Keeping From aligned avoids
        # spoofing signals and improves SPF/DKIM/DMARC deliverability.
        from_email = [smtp_username[0]]
    msg = EmailMessage()
    from_name = _safe_header(settings.get("smtp_from_name") or settings.get("name"), "Maktaba Print")
    msg["From"] = formataddr((from_name, from_email[0]))
    msg["To"] = ", ".join(_csv_emails(",".join(to_emails)))
    cc = _csv_emails(settings.get("report_email_cc", ""))
    bcc = _csv_emails(settings.get("report_email_bcc", ""))
    if cc:
        msg["Cc"] = ", ".join(cc)
    if settings.get("report_email_reply_to"):
        reply_to = _csv_emails(settings["report_email_reply_to"])
        if len(reply_to) != 1:
            raise HTTPException(status_code=422, detail="Adresse Reply-To invalide")
        msg["Reply-To"] = reply_to[0]
    msg["Subject"] = _safe_header(subject, "Rapport Maktaba Print")
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid(domain=from_email[0].split("@", 1)[1])
    msg["Auto-Submitted"] = "auto-generated"
    msg["X-Auto-Response-Suppress"] = "All"
    msg.set_content(text_body or "Votre rapport Maktaba Print est disponible dans la version HTML de cet email.")
    msg.add_alternative(html_body, subtype="html")
    for filename, content in attachments or []:
        msg.add_attachment(
            content,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=_safe_header(filename, "factures.xlsx"),
        )

    timeout = int(settings.get("smtp_timeout_seconds") or 30)
    port = int(settings.get("smtp_port") or 587)
    security = settings.get("smtp_security") or "starttls"
    if security == "ssl" and port == 587:
        raise HTTPException(status_code=400, detail="Le port 587 exige STARTTLS. Utilisez STARTTLS ou le port 465 avec SSL/TLS.")
    if security == "starttls" and port == 465:
        raise HTTPException(status_code=400, detail="Le port 465 exige SSL/TLS. Utilisez SSL/TLS ou le port 587 avec STARTTLS.")
    if settings.get("smtp_username") and not settings.get("smtp_password"):
        raise HTTPException(status_code=400, detail="Mot de passe SMTP non configure. Enregistrez un mot de passe d'application.")
    recipients = list(dict.fromkeys(_csv_emails(",".join(to_emails + cc + bcc))))

    try:
        if security == "ssl":
            with smtplib.SMTP_SSL(settings["smtp_host"], port, timeout=timeout, context=ssl.create_default_context()) as server:
                if settings.get("smtp_username"):
                    server.login(settings["smtp_username"], settings.get("smtp_password", ""))
                server.send_message(msg, to_addrs=recipients)
        else:
            with smtplib.SMTP(settings["smtp_host"], port, timeout=timeout) as server:
                server.ehlo()
                if security == "starttls":
                    server.starttls(context=ssl.create_default_context())
                    server.ehlo()
                if settings.get("smtp_username"):
                    server.login(settings["smtp_username"], settings.get("smtp_password", ""))
                server.send_message(msg, to_addrs=recipients)
    except Exception as exc:
        logger.warning("SMTP delivery failed: %s", type(exc).__name__)
        raise HTTPException(status_code=502, detail=_smtp_error_message(exc))


def _smtp_error_message(exc: Exception) -> str:
    if isinstance(exc, HTTPException):
        return str(exc.detail)
    if isinstance(exc, smtplib.SMTPAuthenticationError):
        return "Authentification SMTP refusee. Verifiez l'identifiant et le mot de passe d'application."
    if isinstance(exc, smtplib.SMTPRecipientsRefused):
        return "Le serveur SMTP a refuse un ou plusieurs destinataires."
    if isinstance(exc, (TimeoutError, smtplib.SMTPServerDisconnected)):
        return "Le serveur SMTP ne repond pas. Verifiez l'hote, le port et le mode de securite."
    if isinstance(exc, ssl.SSLError):
        return "Connexion TLS/SSL impossible. Verifiez le port et le mode de securite SMTP."
    return "Echec de l'envoi. Verifiez la configuration SMTP et la connexion internet."


@router.get("/email/status")
def report_email_status(user=Depends(get_current_user)):
    settings = _load_settings()
    recipients = _csv_emails(settings.get("report_email_recipients", "")) if settings.get("report_email_recipients") else []
    return {
        "enabled": bool(settings.get("report_email_enabled")),
        "smtp_ready": bool(settings.get("smtp_host") and settings.get("smtp_from_email")),
        "password_configured": bool(settings.get("smtp_password")),
        "recipient_count": len(recipients),
        "last_sent_at": settings.get("report_schedule_last_sent_at", ""),
        "last_attempt_at": settings.get("report_schedule_last_attempt_at", ""),
        "last_status": settings.get("report_schedule_last_status", "never"),
        "last_error": settings.get("report_schedule_last_error", ""),
    }

@router.get("/profit")
def profit_report(period: str = "month", db: Session = Depends(get_db), user=Depends(get_current_user)):
    today = date.today()
    if period == "month":
        start = today.replace(day=1)
    elif period == "year":
        start = today.replace(month=1, day=1)
    else:
        start = today

    revenue = db.query(func.sum(Sale.total_amount)).filter(
        Sale.doc_type == "invoice", Sale.status.in_(ACTIVE_SALE_STATUSES),
        func.date(Sale.date_time) >= start,
    ).scalar() or ZERO

    cogs = db.query(func.sum(SaleItem.purchase_price * SaleItem.quantity)).join(Sale).filter(
        Sale.doc_type == "invoice", Sale.status.in_(ACTIVE_SALE_STATUSES),
        func.date(Sale.date_time) >= start,
    ).scalar() or ZERO

    expenses = db.query(func.sum(Expense.amount)).filter(
        func.date(Expense.date) >= start,
    ).scalar() or ZERO

    gross_profit = revenue - cogs
    net_profit = gross_profit - expenses
    margin = (gross_profit / revenue * 100) if revenue > 0 else 0

    return {
        "period": period,
        "revenue": quantize_money(revenue),
        "cogs": quantize_money(cogs),
        "gross_profit": quantize_money(gross_profit),
        "expenses": quantize_money(expenses),
        "net_profit": quantize_money(net_profit),
        "margin_pct": round(margin, 2),
    }


@router.get("/overview")
def reports_overview(
    period: str = "monthly",
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    period_map = {"day": "daily", "week": "weekly", "month": "monthly", "year": "yearly"}
    period_type = period_map.get(period, period)
    start, end = _date_range(period_type, start_date, end_date)
    prev_start, prev_end = _period_comparison_range(start, end)

    summary = _build_summary(db, start, end)
    previous = _build_summary(db, prev_start, prev_end)
    cash = _cash_summary(db, start, end)
    categories = _sales_by_category_range(db, start, end)
    stock = _stock_value_detail(db)
    timeseries = _report_timeseries(db, start, end)

    return {
        "period": {
            "type": period_type,
            "start": start.isoformat(),
            "end": end.isoformat(),
            "previous_start": prev_start.isoformat(),
            "previous_end": prev_end.isoformat(),
        },
        "summary": summary,
        "previous": previous,
        "trend": {
            "revenue": _variation(summary["revenue"], previous["revenue"]),
            "paid": _variation(summary["paid"], previous["paid"]),
            "expenses": _variation(summary["expenses"], previous["expenses"]),
            "purchases": _variation(summary["purchases"], previous["purchases"]),
            "net_profit": _variation(summary["net_profit"], previous["net_profit"]),
        },
        "cash": cash,
        "stock": stock,
        "categories": categories,
        "users": _sales_by_user_range(db, start, end),
        "timeseries": timeseries,
        "top_items": _top_items(db, start, end),
        "dormant_products": _dormant_products(db),
        "hourly_performance": _hourly_performance(db, start, end),
        "comparisons": _business_comparisons(db),
    }


@router.get("/sales-by-category")
def sales_by_category(db: Session = Depends(get_db), user=Depends(get_current_user)):
    today = date.today()
    first_of_month = today.replace(day=1)
    from models.product import Category
    rows = db.query(
        Category.name.label("category"),
        func.sum(SaleItem.line_total).label("total"),
        func.count(SaleItem.id).label("count"),
    ).join(Product, SaleItem.product_id == Product.id
    ).join(Category, Product.category_id == Category.id
    ).join(Sale, SaleItem.sale_id == Sale.id
    ).filter(
        Sale.doc_type == "invoice", Sale.status.in_(ACTIVE_SALE_STATUSES),
        func.date(Sale.date_time) >= first_of_month,
    ).group_by(Category.name).order_by(func.sum(SaleItem.line_total).desc()).all()
    return [{"category": r.category, "total": quantize_money(r.total or 0), "count": r.count} for r in rows]

@router.get("/stock-value")
def stock_value(db: Session = Depends(get_db), user=Depends(get_current_user)):
    from models.product import Category
    rows = db.query(
        Product.id, Product.name, Product.code,
        Product.stock_quantity, Product.purchase_price, Product.sale_price,
        Category.name.label("category"),
    ).join(Category, Product.category_id == Category.id, isouter=True
    ).filter(Product.is_active == 1, Product.product_type == "product"
    ).order_by(func.coalesce(Category.name, "zzz"), Product.name).all()
    total_value = sum((r.stock_quantity or 0) * (r.purchase_price or 0) for r in rows)
    return {
        "total_value": quantize_money(total_value),
        "products": [{
            "id": r.id, "name": r.name, "code": r.code or "",
            "category": r.category or "—",
            "stock": r.stock_quantity, "purchase_price": r.purchase_price,
            "sale_price": r.sale_price,
            "value": quantize_money((r.stock_quantity or 0) * (r.purchase_price or 0)),
        } for r in rows]
    }


@router.post("/email/send")
def send_report_email(body: ReportEmailRequest, db: Session = Depends(get_db), user=Depends(get_current_user)):
    settings = _load_settings()
    start, end = _date_range(body.period_type, body.start_date, body.end_date)
    recipients = _csv_emails(body.recipients) or _csv_emails(settings.get("report_email_recipients", ""))

    data = _build_email_report_data(db, start, end, body)
    subject = body.subject or f"{settings.get('report_email_subject_prefix') or 'Rapport Maktaba Print'} - {start.isoformat()} / {end.isoformat()}"
    html_body = _render_report_html(settings, body, start, end, data)
    excel_bytes, excel_name = _build_invoices_workbook(db, settings, start, end, data)
    _send_email(settings, recipients, subject, html_body, _render_report_text(settings, start, end, data), [(excel_name, excel_bytes)])

    now = datetime.now(ZoneInfo(settings.get("report_schedule_timezone") or "Africa/Casablanca"))
    settings["report_schedule_last_sent_at"] = now.isoformat()
    settings["report_schedule_last_attempt_at"] = now.isoformat()
    settings["report_schedule_last_status"] = "success"
    settings["report_schedule_last_error"] = ""
    _save_settings(settings)
    return {"ok": True, "recipients": recipients, "period": {"start": start.isoformat(), "end": end.isoformat()}, "subject": subject}


@router.post("/email/test")
def test_report_email(body: ReportEmailTestRequest, user=Depends(get_current_user)):
    settings = _load_settings()
    recipient = body.recipient or settings.get("smtp_from_email") or settings.get("report_email_recipients", "")
    recipients = _csv_emails(recipient)
    subject = f"{settings.get('report_email_subject_prefix') or 'Rapport Maktaba Print'} - Test SMTP"
    html_body = "<p>Votre configuration email Maktaba Print fonctionne correctement.</p>"
    _send_email(settings, recipients, subject, html_body, "Votre configuration email Maktaba Print fonctionne correctement.")
    return {"ok": True, "recipients": recipients}
